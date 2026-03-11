"""
Parche: corrige terminología argentina en el JSON y regenera el Excel.
"""
import json, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

JSON_FILE = r'catalogo-organizado.json'
XLSX_FILE = r'catalogo-organizado.xlsx'

# ─── 1. PATCH JSON ───────────────────────────────────────────────────────────

with open(JSON_FILE, 'r', encoding='utf-8') as f:
    data = json.load(f)

# Replacements in all string fields
FIXES = {
    # Categories
    'Molinetes': 'Reels Frontales',
    'Molinete': 'Reel Frontal',
    'Manga Crimping': 'Mangas',
    # In product names
    'Reel Baitcast ': 'Reel ',  # simplify, "Reel TATULA..." reads better than "Reel Baitcast TATULA..."
}

changes = 0
for p in data['productos']:
    for field in ['nombre_es', 'categoria', 'subcategoria']:
        original = p.get(field, '')
        fixed = original
        for old, new in FIXES.items():
            fixed = fixed.replace(old, new)
        if fixed != original:
            p[field] = fixed
            changes += 1

# Fix metadata categories list
data['metadata']['categorias'] = sorted(set(p['categoria'] for p in data['productos']))

# Fix "Reels Baitcast" category → just "Reels Baitcast" is fine, 
# but let's also rename the Molinete in cat order
# Actually the category "Reels Baitcast" the prefix "Reel Baitcast " in nombre_es was simplified to "Reel "

with open(JSON_FILE, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"✅ JSON parcheado: {changes} campos corregidos")

# ─── 2. REGENERAR EXCEL ──────────────────────────────────────────────────────

THIN_BORDER = Border(
    left=Side(style='thin', color='B4C6E7'), right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'), bottom=Side(style='thin', color='B4C6E7'),
)
HEADER_FONT = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
SUBCAT_FONT = Font(name='Calibri', bold=True, size=11, color='2F5496')
SUBCAT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
DATA_FONT   = Font(name='Calibri', size=11)
PRICE_FONT  = Font(name='Calibri', size=11, bold=True)

CAT_COLORS = {
    'Señuelos Artificiales': '2E75B6', 'Cañas': 'BF8F00', 'Reels Frontales': 'C00000',
    'Reels Baitcast': 'C00000', 'Líneas': '7030A0', 'Anzuelos': '548235',
    'Triples': '548235', 'Giradores': '548235', 'Terminal': '548235',
    'Cables de Acero': '404040', 'Accesorios': '404040', 'Indumentaria': '7F6000',
    'Combos': '2E75B6', 'Motores': '404040', 'Repuestos Cañas': '808080',
    'Repuestos': '808080', 'Otros': '808080',
}

CATEGORY_ORDER = [
    'Señuelos Artificiales', 'Cañas', 'Reels Frontales', 'Reels Baitcast',
    'Líneas', 'Anzuelos', 'Triples', 'Giradores', 'Terminal',
    'Cables de Acero', 'Accesorios', 'Indumentaria', 'Combos',
    'Motores', 'Repuestos Cañas', 'Repuestos', 'Otros',
]

def safe_name(n):
    n = n[:31]
    for c in '\\/*?:[]': n = n.replace(c, '')
    return n

def write_sheet(ws, items, color):
    hfill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    headers = ['CÓDIGO', 'PRODUCTO', 'SUBCATEGORÍA', 'MARCA', 'PRECIO (USD)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT; c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center'); c.border = THIN_BORDER
    ws.freeze_panes = 'A2'
    items.sort(key=lambda x: (x.get('subcategoria',''), x.get('nombre_es','')))
    cur_sub = None; row = 2
    for item in items:
        sub = item.get('subcategoria', '')
        if sub != cur_sub:
            cur_sub = sub
            for col in range(1, 6):
                c = ws.cell(row=row, column=col)
                if col == 1: c.value = f"▸ {cur_sub}"
                c.font = SUBCAT_FONT; c.fill = SUBCAT_FILL; c.border = THIN_BORDER
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1
        ws.cell(row=row, column=1, value=item['codigo']).font = DATA_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=item['nombre_es']).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=item.get('subcategoria','')).font = DATA_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=4, value=item.get('marca','')).font = DATA_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER
        pc = ws.cell(row=row, column=5, value=item['precio_usd'])
        pc.font = PRICE_FONT; pc.number_format = '#,##0.00'
        pc.border = THIN_BORDER; pc.alignment = Alignment(horizontal='right')
        row += 1
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 62
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.auto_filter.ref = f'A1:E{row - 1}'

def write_index(ws, cat_counts):
    ws.sheet_properties.tabColor = '2F5496'
    ws.merge_cells('A1:D1')
    t = ws.cell(row=1, column=1, value='CATÁLOGO DE PRODUCTOS')
    t.font = Font(name='Calibri', bold=True, size=18, color='2F5496')
    t.alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:D2')
    s = ws.cell(row=2, column=1, value='Lista de Precios Mayorista (USD)')
    s.font = Font(name='Calibri', size=13, color='808080', italic=True)
    s.alignment = Alignment(horizontal='center')
    row = 4
    for col, h in enumerate(['CATEGORÍA', 'PRODUCTOS', 'PESTAÑA', 'PRECIO PROM.'], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        c.alignment = Alignment(horizontal='center'); c.border = THIN_BORDER
    row = 5; total = 0
    for cat in CATEGORY_ORDER:
        if cat not in cat_counts: continue
        info = cat_counts[cat]
        total += info['count']
        color = CAT_COLORS.get(cat, '2F5496')
        cfill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        ws.cell(row=row, column=1, value=cat).font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        ws.cell(row=row, column=1).fill = cfill; ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=info['count']).font = DATA_FONT
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center'); ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=safe_name(cat)).font = Font(name='Calibri', size=11, color='2F5496', underline='single')
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center'); ws.cell(row=row, column=3).border = THIN_BORDER
        avg = info['total_price'] / info['count'] if info['count'] > 0 else 0
        pc = ws.cell(row=row, column=4, value=avg)
        pc.font = DATA_FONT; pc.number_format = '#,##0.00'; pc.alignment = Alignment(horizontal='right'); pc.border = THIN_BORDER
        row += 1
    ws.cell(row=row, column=1, value='TOTAL').font = Font(name='Calibri', bold=True, size=12)
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=total).font = Font(name='Calibri', bold=True, size=12)
    ws.cell(row=row, column=2).alignment = Alignment(horizontal='center'); ws.cell(row=row, column=2).border = THIN_BORDER
    ws.column_dimensions['A'].width = 28; ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 28; ws.column_dimensions['D'].width = 16

products = data['productos']
by_cat = defaultdict(list)
cat_stats = {}
for p in products:
    cat = p['categoria']
    by_cat[cat].append(p)
    if cat not in cat_stats: cat_stats[cat] = {'count': 0, 'total_price': 0}
    cat_stats[cat]['count'] += 1
    cat_stats[cat]['total_price'] += p.get('precio_usd', 0)

wb = Workbook()
ws_idx = wb.active; ws_idx.title = 'ÍNDICE'
for cat in CATEGORY_ORDER:
    if cat not in by_cat: continue
    sn = safe_name(cat)
    ws = wb.create_sheet(title=sn)
    ws.sheet_properties.tabColor = CAT_COLORS.get(cat, '2F5496')
    write_sheet(ws, by_cat[cat], CAT_COLORS.get(cat, '2F5496'))
    print(f"  ✓ {cat}: {len(by_cat[cat])} productos")

write_index(ws_idx, cat_stats)
wb.save(XLSX_FILE)
print(f"\n✅ Excel regenerado: {XLSX_FILE}")
